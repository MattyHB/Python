import sys
import openpyxl
skip = ''
try:
    filename = sys.argv[1]
    colnum = sys.argv[2]
except Exception as e:
    print('Usage: dumpxl.py filename.xlsx column-number')
    print('Optional: "-skip"  flag to skip the first line in the file')
    exit(1)
try:
    skip = sys.argv[3]
except Exception as e:
    pass

wb = openpyxl.load_workbook(filename)
sheet = wb.active

if skip == '-skip':
    for row in sheet.iter_cols(min_row=2, min_col =int(colnum), max_col= int(colnum)):
        for cell in row:
            print(cell.value)
else:
    for row in sheet.iter_cols(min_col =int(colnum), max_col= int(colnum)):
        for cell in row:
            print(cell.value)
